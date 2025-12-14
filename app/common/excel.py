from app.common.chardecode import convert
from openpyxl.styles import Font, Border, PatternFill, Alignment, Side, Color
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


class ExcelExport:

    class ColumnInfo:
        def __init__(self, col_letter, title, align='center', decode_yn=False, code_dict=None):
            self.__col_letter = col_letter
            self.__title = title
            self.__align = align
            self.__decode_yn = decode_yn
            self.__code_dict = code_dict

        @property
        def col_letter(self):
            return self.__col_letter

        @property
        def title(self):
            return self.__title

        @property
        def align(self):
            return self.__align

        @property
        def decode_yn(self):
            return self.__decode_yn

        @property
        def code_dict(self):
            return self.__code_dict

    def __init__(self):
        self._column_definition = {}

    def keys(self):
        return self._column_definition.keys()

    def get(self, key):
        return self._column_definition[key]

    def set_header_column(self, worksheet, start_row=3, fixed_row=None):
        for k, v in self._column_definition.items():
            cell = v.col_letter + str(start_row)
            worksheet[cell] = v.title
            worksheet[cell].font = Font(bold=True)
            worksheet[cell].border = Border(left=Side(border_style='thin', color='000000'),
                                            right=Side(border_style='thin', color='000000'),
                                            top=Side(border_style='thin', color='000000'),
                                            bottom=Side(border_style='thin', color='000000'))
            worksheet[cell].fill = PatternFill(patternType='solid', fgColor=Color("D9D9D9"))
            worksheet[cell].alignment = Alignment(horizontal='center', vertical='center')

        if fixed_row is not None:
            worksheet.freeze_panes = 'A' + str(fixed_row)

    def set_data_column(self, worksheet, data_row, start_row=4):
        # col loop
        for k, v in self._column_definition.items():
            cell = v.col_letter + str(start_row)
            rest_val = data_row.get(k)

            # cell format
            worksheet[cell].alignment = Alignment(horizontal=v.align, vertical='center', wrap_text=True)
            worksheet[cell].border = Border(left=Side(border_style='thin', color='000000'),
                                            right=Side(border_style='thin', color='000000'),
                                            top=Side(border_style='thin', color='000000'),
                                            bottom=Side(border_style='thin', color='000000'))

            # cell value
            if rest_val is None:
                continue

            if v.decode_yn:  # Export 시 Decode Base64 처리
                # ws[cell] = base64.b64decode(rest_val.encode()).decode().replace("\n", "\r\n")

                # Decode Base64
                converttext = convert(textinput = rest_val)
                worksheet[cell] = ILLEGAL_CHARACTERS_RE.sub(r'?', converttext[0])
            else:
                worksheet[cell] = rest_val

                if v.code_dict is not None:
                    worksheet[cell] = v.code_dict[rest_val]
                else:
                    worksheet[cell] = rest_val

    @staticmethod
    def set_additional_info_row(worksheet, additional_info, cell="B2"):
        worksheet[cell] = additional_info
        worksheet[cell].font = Font(color='000000', bold=False, underline='none', size=12)

    # instance.autofit_cell_size(ws, [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11])
    # instance.autofit_cell_size(ws)
    @staticmethod
    def autofit_cell_size(worksheet, target_cols=None, margin=2):
        # i is col
        for i, col_cells in enumerate(worksheet.columns):
            is_ok = False
            max_length = 0

            if target_cols is None:
                is_ok = True
            elif isinstance(target_cols, list) and i in target_cols:
                is_ok = True

            if is_ok:
                for cell in col_cells:
                    # max() 함수 호출 전 시퀀스가 비어있는지 확인하여 ValueError 방지
                    sequence = str(cell.value).splitlines()
                    encode_sequence = str(cell.value).encode().splitlines()

                    if sequence and encode_sequence:
                        encode_length = len(max(encode_sequence, key=len))
                        normal_length = len(max(sequence, key=len))
                    else:
                        encode_length = 0
                        normal_length = 0

                    length = ((normal_length + encode_length) / 2) if normal_length != encode_length else normal_length

                    if length > max_length:
                        max_length = length
                worksheet.column_dimensions[col_cells[0].column_letter].width = max_length + margin
                # worksheet.column_dimensions[col_cells[0].column_letter].auto_size = True

        # row size
        for i, row_cells in enumerate(worksheet.rows):
            # worksheet.row_dimensions[row_cells[0].row].auto_size = True
            worksheet.row_dimensions[row_cells[0].row].height = 16.5
