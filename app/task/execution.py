import base64
import os
import json
from datetime import datetime
from queue import Queue

from app.common.chardecode import convert
from flask import Blueprint, jsonify, request, current_app, send_file, Response, stream_with_context
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import PatternFill
from werkzeug.exceptions import abort

from app.common.code import CommonCode
from app.common.excel import ExcelExport
from app.common.render import Render
from app.interface.constants import OpmmResultCode
from app.interface.restclient import RestClient
from app.user import login_required

bp = Blueprint('execution', __name__)
sse_dict = {}


@bp.route('/', methods=['GET'])
@login_required
def render_page():
    """
    태스크실행 목록 화면
    :return:
    """
    return Render.render_template('task/execution_lst.html', sse_enable=current_app.config['SSE_ENABLE'])


@bp.route('/', methods=['POST'])
@login_required
def render_page_with_param():
    """
    태스크실행 목록 화면
    :return:
    """
    return Render.render_template('task/execution_lst.html', request_params=request.form, sse_enable=current_app.config['SSE_ENABLE'])


@bp.route('/list', methods=['POST'])
@login_required
def request_exec_list():
    """
    태스크실행 목록 요청
    :return:
    """
    # Step 1. Request Data Parsing.
    req = request.get_json()

    task_id = req['task_id']
    owner_id = req['owner_id']
    offset = (req['page'] - 1) * req['perPage']
    limit = req['perPage']

    # Step 2. Request Data Validation.
    #   - None.

    # Step 3. REST Call
    rest_execution = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/executions'
    query_string = ''

    if task_id != '':
        query_string = ',taskId=' + task_id
    if owner_id != '':
        query_string += ',taskOwner=' + owner_id

    # field 확장 - permMode 추가 (opmm 230712)
    rest_execution.req = {'query': query_string[1:],  # 좌측 ',' 제거
                          'field': 'permMode',
                          'limit': limit,
                          'offset': offset}

    # Step 3-2. REST Call : get
    if not rest_execution.get(url):
        return jsonify({'result': 'Fail'}), 500

    # Step 3-3. REST Call : Check Biz. Exception
    # NOT_FOUND
    if rest_execution.res['resultCode'] == OpmmResultCode.EM1002:
        rest_execution.res['executionList'] = []
        rest_execution.res['totalCnt'] = 0
        return jsonify(rest_execution.res)

    # Step 3-4. REST Call : Output Process

    # Step 99. Response Data Formatting
    return jsonify(rest_execution.res)


@bp.route('/info', methods=['POST'])
@login_required
def request_exec_info_list():
    """
    태스크실행 단건조회 요청
    :return:
    """
    # Step 1. Request Data Parsing.
    req = request.get_json()

    execution_no = str(req['execution_no'])

    # Step 2. Request Data Validation.
    #   - None.

    # Step 3. REST Call
    rest_execution = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/executions/' + execution_no

    # Step 3-2. REST Call : get
    if not rest_execution.get(url):
        return jsonify({'result': 'Fail'}), 500

    # Step 3-3. REST Call : Check Biz. Exception
    # NOT_FOUND
    if rest_execution.res['resultCode'] != OpmmResultCode.EM0000:
        current_app.logger.error("[{0}] {1}".format(rest_execution.res['resultCode'], rest_execution.res['resultMsg']))

    # Step 3-4. REST Call : Output Process

    # Step 99. Response Data Formatting
    return jsonify(rest_execution.res)


@bp.route('/nodelist', methods=['POST'])
@login_required
def request_exec_nodelist():
    """
    태스크실행 노드목록 요청
    :return:
    """
    # Step 1. Request Data Parsing.
    req = request.get_json()

    execution_no = str(req['execution_no'])
    node_hostname = req.get('node_hostname')
    node_status = None
    node_result = None

    offset = (req['page'] - 1) * req['perPage']
    limit = req['perPage']

    if req.get('node_status') != 'all':
        node_status = req.get('node_status')

    if req.get('node_result') != 'all':
        node_result = req.get('node_result')

    # Step 2. Request Data Validation.
    #   - None.

    # Step 3. REST Call
    rest_execution = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/executions/' + execution_no + '/nodes'
    query_string = ''

    if node_hostname != '' and node_hostname is not None:
        query_string = ',hostname=' + str(node_hostname)
    if node_status != '' and node_status is not None:
        query_string += ',status=' + str(node_status)
    if node_result != '' and node_result is not None:
        query_string += ',result=' + str(node_result)

    rest_execution.req = {'query': query_string[1:],  # 좌측 ',' 제거
                          'limit': limit,
                          'offset': offset}

    # Step 3-2. REST Call : get
    if not rest_execution.get(url):
        return jsonify({'result': 'Fail'}), 500

    # Step 3-3. REST Call : Check Biz. Exception
    # NOT_FOUND
    if rest_execution.res['resultCode'] == OpmmResultCode.EM1002:
        rest_execution.res['executionNodeList'] = []
        rest_execution.res['totalCnt'] = 0

    # Step 3-4. REST Call : Output Process

    # Step 99. Response Data Formatting
    return jsonify(rest_execution.res)


@bp.route('/nodeinfo', methods=['POST'])
@login_required
def request_exec_info_nodelist():
    """
    태스크실행 노드 단건조회 요청
    :return:
    """
    # Step 1. Request Data Parsing.
    req = request.get_json()

    execution_no = str(req['executionNo'])
    nodesession_id = req['nodeSessionId']

    # Step 2. Request Data Validation.
    #   - None.

    # Step 3. REST Call
    rest_execution = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/executions/' + execution_no + '/nodes/' + nodesession_id

    # Step 3-2. REST Call : get
    if not rest_execution.get(url):
        return jsonify({'result': 'Fail'}), 500

    # Step 3-3. REST Call : Check Biz. Exception
    # NOT_FOUND
    if rest_execution.res['resultCode'] != OpmmResultCode.EM0000:
        current_app.logger.error("[{0}] {1}".format(rest_execution.res['resultCode'], rest_execution.res['resultMsg']))

    # Step 3-4. REST Call : Output Process
    # Decode Base64
    stdout_barr = convert(rest_execution.res['base64Stdout'])
    stderr_barr = convert(rest_execution.res['base64Stderr'])

    # Bind base64 encoded string
    rest_execution.res['base64Stdout'] = base64.b64encode(stdout_barr[0].encode()).decode()
    rest_execution.res['base64Stderr'] = base64.b64encode(stderr_barr[0].encode()).decode()

    # Step 99. Response Data Formatting
    return jsonify(rest_execution.res)


@bp.route('/target_list', methods=['POST'])
@login_required
def request_task_target_list():
    """
    태스크실행 태스크 타겟 노드 목록(모의 실행)
    - 태스크의 타겟 노드셋에 해당하는 현재 노드 목록을 조회함.
    - 페이징 없이 해당 기능을 수행하는 OPMM 의 REST API 가 별도로 존재하지 않기에, OPMM 의 Dry Run 사용.
    :return:
    """
    # Step 1. Request Data Parsing.
    req = request.get_json()

    task_id = req['task_id']
    rev_no = req['rev_no']

    # Step 2. Request Data Validation.
    #   - None.

    # Step 3. REST Call
    rest_execution = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/executions?mode=dry'
    rest_execution.req = {'taskId': task_id,
                          'revNo': rev_no}

    # Step 3-2. REST Call : post
    if not rest_execution.post(url):
        return jsonify({'result': 'Fail'}), 500

    # Step 3-3. REST Call : Check Biz. Exception

    # Step 3-4. REST Call : Output Process

    # Step 99. Response Data Formatting
    return jsonify(rest_execution.res)


@bp.route('/dry_run', methods=['POST'])
@login_required
def request_task_dry_run():
    """
    태스크실행 모의 실행
    :return:
    """
    # Step 1. Request Data Parsing.
    req = request.get_json()

    task_id = req['task_id']
    rev_no = req['rev_no']
    target_list = req['target_list']
    # task_arg = req['task_arg']  # 현재는 OPMM 에서 유의미한 처리를 하지 않음.
    runner_type = req['runner_type']
    runner_id = req['runner_id']
    run_date = req['run_date']

    # Step 2. Request Data Validation.
    #   - None.

    # Step 3. REST Call
    rest_execution = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/executions?mode=dry'
    rest_execution.req = {'taskId': task_id,
                          'revNo': rev_no,
                          'nodeSessionIdList': target_list,
                          'runnerType': runner_type,
                          'runnerId': runner_id,
                          'runDate': run_date}

    # Step 3-2. REST Call : post
    if not rest_execution.post(url):
        return jsonify({'result': 'Fail'}), 500

    # Step 3-3. REST Call : Check Biz. Exception

    # Step 3-4. REST Call : Output Process

    # Step 99. Response Data Formatting
    return jsonify(rest_execution.res)


@bp.route('/execute', methods=['POST'])
@login_required
def request_execute():
    """
    태스크실행 실행
    :return:
    """
    # Step 1. Request Data Parsing.
    req = request.get_json()

    task_id = req['task_id']
    target_list = req['target_list']
    task_arg = req['task_arg']

    # Step 2. Request Data Validation.
    #   - None.

    # Step 3. REST Call
    rest_execution = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/executions'

    rest_execution.req = {'taskId': task_id,
                          'nodeSessionIdList': target_list,
                          'argument': task_arg}

    # Step 3-2. REST Call : post
    if not rest_execution.post(url):
        return jsonify({'result': 'Fail'}), 500

    # Step 3-3. REST Call : Check Biz. Exception
    if rest_execution.res['resultCode'] != OpmmResultCode.EM0000:
        current_app.logger.error("[{0}] {1}".format(rest_execution.res['resultCode'], rest_execution.res['resultMsg']))

    # Step 3-4. REST Call : Output Process

    # Step 99. Response Data Formatting
    return jsonify(rest_execution.res)


@bp.route('/force_stop', methods=['POST'])
@login_required
def request_exec_force_stop():
    """
    태스크실행 강제종료
    :return:
    """
    # Step 1. Request Data Parsing.
    req = request.get_json()
    execution_no = req['execution_no']

    # Step 2. Request Data Validation.
    #   - None.

    # Step 3. REST Call(Create&Edit Usergrp)
    rest_execution = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/executions/' + str(execution_no)

    rest_execution.req = {'forceStop': 'Q'}

    # Step 3-2. REST Call : get
    result = rest_execution.put(url)

    if not result:
        return jsonify({'result': 'Fail'}), 500

    # Step 3-3. REST Call : Check Biz. Exception
    if rest_execution.res['resultCode'] != OpmmResultCode.EM0000:
        current_app.logger.error("[{0}] {1}".format(rest_execution.res['resultCode'], rest_execution.res['resultMsg']))

    # Step 3-4. REST Call : Output Process

    # Step 99. Response Data Formatting
    return jsonify(rest_execution.res)


@bp.route('/node_force_stop', methods=['POST'])
@login_required
def request_exec_node_force_stop():
    """
    태스크실행 노드강제종료
    :return:
    """
    # Step 1. Request Data Parsing.
    req = request.get_json()
    execution_no = req['execution_no']
    node_session_id = req['node_session_id']

    # Step 2. Request Data Validation.
    #   - None.

    # Step 3. REST Call(Create&Edit Usergrp)
    rest_execution = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/executions/' + str(execution_no) + '/nodes/' + node_session_id

    rest_execution.req = {'forceStop': 'Q'}

    # Step 3-2. REST Call : get
    result = rest_execution.put(url)

    if not result:
        return jsonify({'result': 'Fail'}), 500

    # Step 3-3. REST Call : Check Biz. Exception
    if rest_execution.res['resultCode'] != OpmmResultCode.EM0000:
        current_app.logger.error("[{0}] {1}".format(rest_execution.res['resultCode'], rest_execution.res['resultMsg']))

    # Step 3-4. REST Call : Output Process

    # Step 99. Response Data Formatting
    return jsonify(rest_execution.res)


@bp.route('/p_node', methods=['GET'])
@login_required
def render_node_dtl_page():
    """
    노드별 태스크 실행 상세 Popup 화면 (단건조회)
    :return:
    """
    return Render.render_template('task/popup/popup_execution_node_dtl.html')


@bp.route('/p_condition', methods=['GET'])
@login_required
def render_condition_exec_page():
    """
    조건실행 Popup 화면
    :return:
    """
    return Render.render_template('task/popup/popup_execution_condition.html')


@bp.route('/p_dry_run', methods=['GET'])
@login_required
def render_dry_run_page():
    """
    모의실행 Popup 화면
    :return:
    """
    return Render.render_template('task/popup/popup_execution_dry_run.html')


@bp.route('/xl_export', methods=['POST'])
@login_required
def request_exec_xl_export():
    """
    실행결과 엑셀 내보내기
    :return:
    """
    # Step 1. Request Data Parsing.
    execution_no = request.form.get('execution_no')
    current_date = datetime.now().strftime("%Y%m%d")
    export_path = current_app.config['OPME_FILE_EXPORT_PATH'] + '/' + current_date

    # 실행 정보 조회 조건 받기
    node_hostname = request.form.get('sc_node_hostname')
    node_status = request.form.get('sc_node_status')
    node_result = request.form.get('sc_node_result')
    save_filename = request.form.get('uuid')

    # Create File Download Directory(from OPMM)
    if os.path.isdir(export_path) is False:
        os.mkdir(export_path)

    # 반복 조회 수행 시 offset, limit 필요.
    offset = 0
    limit = 100
    total_cnt = int(request.form.get('total_cnt'))

    sse_enable = current_app.config['SSE_ENABLE']

    # Step 2. Request Data Validation.
    #   - None.
    class ExecutionExcelExport(ExcelExport):

        def __init__(self):
            super().__init__()

            # Code dictionary
            self.__status_dict = {item['value']: item['text'] for item in getattr(CommonCode, 'execution_node_status')}
            self.__result_dict = {item['value']: item['text'] for item in getattr(CommonCode, 'execution_result')}
            self.__force_stop_dict = {item['value']: item['text'] for item in getattr(CommonCode, 'execution_force_stop')}
            self._column_definition = {
                # key: ColumnInfo(col_letter, title, align, decode_yn, code_dict)
                # Col 'A' : Empty
                'nodeSessionId': self.ColumnInfo('B', '노드세션ID', align='left'),
                'hostname': self.ColumnInfo('C', 'Hostname'),
                'remoteAddr': self.ColumnInfo('D', 'IP주소'),
                'account': self.ColumnInfo('E', '실행계정'),
                'status': self.ColumnInfo('F', '상태', code_dict=self.__status_dict),
                'result': self.ColumnInfo('G', '결과', code_dict=self.__result_dict),
                'exitNum': self.ColumnInfo('H', '실행종료코드'),
                'forceStop': self.ColumnInfo('I', '강제종료', code_dict=self.__force_stop_dict),
                'stopCause': self.ColumnInfo('J', '종료사유'),
                'startDate': self.ColumnInfo('K', '시작일시'),
                'endDate': self.ColumnInfo('L', '종료일시'),
                'base64Stdout': self.ColumnInfo('M', '표준출력', align='left', decode_yn=True),
                'base64Stderr': self.ColumnInfo('N', '표준에러', align='left', decode_yn=True),
            }

    # OpenPyXL
    wb = Workbook()
    ws = wb.active
    ws.title = "OPMATE"
    ex = ExecutionExcelExport()

    # Set table header columns
    ex.set_header_column(ws, fixed_row=4)

    # Step 3. REST Call
    rest_execution = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/executions/' + str(execution_no) + '/nodes'
    field_string = 'account,exitNum,forceStop,stopCause,startDate,endDate,base64Stdout,base64Stderr'
    query_string = ''

    if node_hostname != '' and node_hostname is not None:
        query_string = ',hostname=' + str(node_hostname)
    if node_status != '' and node_status is not None:
        query_string += ',status=' + str(node_status)
    if node_result != '' and node_result is not None:
        query_string += ',result=' + str(node_result)

    rest_execution.req = {'query': query_string[1:],  # 좌측 ',' 제거
                          'field': field_string,
                          'limit': limit,
                          'offset': offset}

    current_cnt = 0
    while total_cnt > offset:

        rest_execution.req['offset'] = offset

        # Step 3-2. REST Call : get
        if not rest_execution.get(url):
            return jsonify({'result': 'Fail'}), 500

        # Step 3-3. REST Call : Check Biz. Exception
        if rest_execution.res['resultCode'] != OpmmResultCode.EM0000:
            current_app.logger.error(
                "[{0}] {1}".format(rest_execution.res['resultCode'], rest_execution.res['resultMsg']))
            abort(500)

        # row loop
        for n in range(rest_execution.res['currentCnt']):  # totalCnt 15 면 range(0, 10), range(0, 5) 으로 실행
            ex.set_data_column(ws, rest_execution.res['executionNodeList'][n], n + offset + 4)

            # Apply text style
            for col_idx in range(1, ws.max_column + 1):
                cell_ref = ws.cell(row=n + offset + 4, column=col_idx)
                cell_value = cell_ref.value

                # Add single quote - stdout, stderr
                if col_idx in [13, 14]:
                    cell_ref.value = f"'{cell_value}"
                cell_ref.number_format = '@'

            current_cnt += 1

            if sse_enable == 'yes':
                if sse_dict.get(save_filename) is None:
                    current_app.logger.error("Aborted excel file export.")
                    delete_file = os.path.join(export_path, save_filename)
                    try:
                        if os.path.isfile(delete_file):
                            os.remove(delete_file)
                    except OSError:
                        abort(500)
                    abort(500)

                sse_dict[save_filename].put(u'%s' % int(current_cnt / total_cnt * 100))

        offset += limit

    # 열 너비 맞춤
    ex.autofit_cell_size(ws)

    # Last Summary add
    summary = "태스크ID : " + str(request.form.get('task_id')) + \
              ", 실행ID : " + str(request.form.get('runner_id')) + \
              ", 시작일시 : " + str(request.form.get('start_date')) + \
              ", 종료일시 : " + str(request.form.get('end_date')) + \
              ", 총 건수 : " + str(total_cnt)
    ex.set_additional_info_row(ws, summary)

    # 파일명 지정, 저장 처리
    file_name = export_path + "/" + save_filename
    wb.save(filename=file_name)

    '''
    # 처리 변경 (send_file 아닌 방법)
    def generate():
        try:
            with open(file_name, "rb") as f:
                yield from f
        except Exception as e:
            print("Exception : {0}".format(e))
        print('delete_file start')
        os.remove(file_name)
        print('delete_file end')
    r = current_app.response_class(generate(), mimetype='application/octet-stream')
    r.headers.set('Content-Disposition', 'attachment', filename=file_name)
    return r
    '''
    return send_file(os.path.join(export_path, file_name),
                     attachment_filename=current_date + '/' + file_name,
                     as_attachment=True)


@bp.route('/gather_export', methods=['POST'])
@login_required
def request_exec_gather_export():
    """
    실행결과 gather 엑셀 내보내기
    :return:
    """
    # Step 1. Request Data Parsing.
    # execution_no = request.form.get('execution_no')
    current_date = datetime.now().strftime("%Y%m%d")
    export_path = current_app.config['OPME_FILE_EXPORT_PATH'] + '/' + current_date
    save_filename = request.json.get('uuid')

    # Create File Download Directory(from OPMM)
    if os.path.isdir(export_path) is False:
        os.mkdir(export_path)

    data = request.json

    # gather 디코딩
    decoded_data = json.loads(base64.b64decode(data['gather_data']).decode('utf-8'))
    sse_enable = current_app.config['SSE_ENABLE']
    # OpenPyXL
    wb = Workbook()
    ws = wb.active
    ws.title = "OPMATE"

    # 헤더
    headers = ["노드세션ID", "Hostname", "IP주소", "결과"]
    gather_keys = set()

    # result 값만 치환.
    def get_result_text(value):
        for items in CommonCode.execution_result:
            if items['value'] == value:
                return items['text']
        return value

    rows = []
    for item in decoded_data:
        col1 = item.get("nodeSessionId", "")
        col2 = item.get("hostname", "")
        col3 = item.get("remoteAddr", "")
        col4 = get_result_text(item.get("result", ""))

        row = [col1, col2, col3, col4]

        gather = item.get("gather", {})
        for key in gather.keys():
            if key not in gather_keys:
                gather_keys.add(key)
                headers.append(key)
            row.append(gather.get(key, ""))
        rows.append(row)
    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for row in rows:
        ws.append(row)
        if sse_enable == 'yes':
            sse_dict[save_filename].put(u'%s' % int(len(row) / len(rows) * 100))

    # 열 너비 자동 맞춤 함수
    def autofit_cell_size(worksheet):
        for cols in worksheet.columns:
            max_length = 0
            column = cols[0].column_letter  # Get the column name
            for cells in cols:
                try:
                    if len(str(cells.value)) > max_length:
                        max_length = len(cells.value)
                finally:
                    pass
            adjusted_width = (max_length + 4)
            worksheet.column_dimensions[column].width = adjusted_width

    # 열 너비 자동 맞춤 적용
    autofit_cell_size(ws)

    file_name = export_path + "/" + save_filename
    wb.save(filename=file_name)

    return send_file(os.path.join(export_path, file_name), download_name=current_date + '/' + file_name, as_attachment=True)


@bp.route('/del_tmp_file', methods=['POST'])
@login_required
def delete_tmp_file():
    """
    실행결과 엑셀 내보내기 완료 후, 호출하여 임시 파일 삭제.
    :return:
    """
    req = request.get_json()
    date = req['date']
    file_name = req['file_name']
    export_path = current_app.config['OPME_FILE_EXPORT_PATH'] + "/" + date
    delete_file = os.path.join(export_path, file_name)

    try:
        if os.path.isfile(delete_file):
            os.remove(delete_file)

    except OSError:
        return jsonify({'result': 'Fail'}), 500

    return jsonify({'result': 'Success'})


@bp.route('/sse/<uuid>', methods=['GET'])
@login_required
def export_sse(uuid):
    if not sse_dict.get(uuid) is None:
        return None
    return Response(stream_with_context(event_stream(uuid)),
                    mimetype="text/event-stream",
                    headers={"X-Accel-Buffering": "no"})


def event_stream(uuid):
    try:
        sse_dict[uuid] = Queue()
        progress_percent = 0

        while int(progress_percent) < 100:
            progress_percent = sse_dict[uuid].get()
            yield 'data: %s\n\n' % progress_percent
    finally:
        if not sse_dict.get(uuid) is None:
            del(sse_dict[uuid])
