from flask import Blueprint, jsonify, request, current_app, send_file, Response, stream_with_context

import os
import json
from datetime import datetime
from queue import Queue

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import PatternFill
from werkzeug.exceptions import abort

from app.common.chardecode import convert
from app.common.code import CommonCode
from app.common.excel import ExcelExport
from app.common.render import Render
from app.interface.constants import OpmmResultCode
from app.interface.restclient import RestClient
from app.user import login_required

bp = Blueprint('node', __name__)
sse_dict = {}

def _parse_json_field(s):
    try:
        return json.loads(s) if s else {}
    except Exception:
        return {}

def get_node_list_data(query_string: str = "", 
                       regexp_string: str = "", 
                       limit: int = 20, 
                       offset: int = 0, 
                       field_string: str = "") -> dict:
    """
    노드 목록 데이터 조회 공통 함수
    
    Args:
        query_string: 쿼리 문자열
        regexp_string: 정규식 검색 문자열 (nodeSet)
        limit: 조회 개수
        offset: 시작 위치
        field_string: 추가 필드
        
    Returns:
        dict: 노드 목록 데이터
        
    Note:
        이 함수는 Flask 애플리케이션 컨텍스트 내에서 호출되어야 합니다.
        RestClient()가 current_app.config를 사용하기 때문입니다.
    """
    rest_node = RestClient()
    
    url = '/nodes'
    rest_node.req = {
        'query': query_string,
        'nodeSet': regexp_string,
        'limit': limit,
        'offset': offset,
        'field': field_string
    }
    
    # REST Call
    if not rest_node.get(url):
        return {'result': 'Fail', 'nodeList': [], 'totalCnt': 0}
    
    # Check Biz. Exception - NOT_FOUND
    if rest_node.res['resultCode'] == OpmmResultCode.EM1002:
        rest_node.res['nodeList'] = []
        rest_node.res['totalCnt'] = 0
    
    return rest_node.res


@bp.route('/', methods=['GET'])
@login_required
def render_page():
    """
    노드 목록 화면
    :return:
    """
    return Render.render_template('node/node_lst.html')


@bp.route('/', methods=['POST'])
@login_required
def render_page_with_param():
    """
    노드 목록 화면
    :return:
    """
    return Render.render_template('node/node_lst.html', request_params=request.form)


@bp.route('/list', methods=['POST'])
@login_required
def request_node_list():
    """
    노드 목록 요청
    :return:
    """
    # Step 1. Request Data Parsing.
    req = request.get_json()

    offset = (req['page'] - 1) * req['perPage']
    limit = req['perPage']

    # Step 2. Request Data Validation.
    #   - None.

    # Step 3. REST Call
    # Step 3-1. REST Call : Input Binding.
    query_string = ""
    regexp_string = ""
    field_string = ""

    if not req.get('base_condition') is None:
        for key, val in req.get('base_condition').items():
            if not (val is None or val == ""):
                query_string += ',' + key + '=' + val

    if not req.get('nodeset_svr_condition') is None:
        for key, val in req.get('nodeset_svr_condition').items():
            if not (val is None or val == ""):
                # regexp_string += ',' + key + ':\"' + val + '\"'
                regexp_string += ',' + key + ':\"' + val + '\"'

    # CT.Name:"ec2.*",CT.Creator:"Terraform"
    if not req.get('nodeset_tag_condition') is None:
        for key, val in req.get('nodeset_tag_condition').items():
            if not (val is None or val == ""):
                regexp_string += ',' + key + ':\"' + val + '\"'

    if not req.get('extra_fields') is None:
        field_string = req.get('extra_fields')

    # 좌측 ',' 제거
    query_string = query_string[1:] if query_string else ""
    regexp_string = regexp_string[1:] if regexp_string else ""

    # for TEST
    current_app.logger.debug(f"query: {query_string}, nodeSet: {regexp_string}")

    # 공통 함수 호출
    result = get_node_list_data(
        query_string=query_string,
        regexp_string=regexp_string,
        limit=limit,
        offset=offset,
        field_string=field_string
    )

    # Step 99. Response Data Formatting
    return jsonify(result)


@bp.route('/list_all', methods=['POST'])
@login_required
def request_node_list_all():
    """
    노드 목록 요청
    :return:
    """
    # Step 1. Request Data Parsing.
    req = request.get_json()

    # Step 2. Request Data Validation.
    #   - None.

    # Step 3. REST Call
    rest_node = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/nodes'
    query_string = ""
    regexp_string = ""
    field_string = ""

    if not req.get('base_condition') is None:
        for key, val in req.get('base_condition').items():
            if not (val is None or val == ""):
                query_string += ',' + key + '=' + val

    if not req.get('nodeset_svr_condition') is None:
        for key, val in req.get('nodeset_svr_condition').items():
            if not (val is None or val == ""):
                regexp_string += ',' + key + ':\"' + val + '\"'

    # CT.Name:"ec2.*",CT.Creator:"Terraform"
    if not req.get('nodeset_tag_condition') is None:
        for key, val in req.get('nodeset_tag_condition').items():
            if not (val is None or val == ""):
                regexp_string += ',' + key + ':\"' + val + '\"'

    if not req.get('extra_fields') is None:
        field_string = req.get('extra_fields')

    offset = 0
    limit = 1000
    node_list = []

    rest_node.req = {'query': query_string[1:],  # 좌측 ',' 제거
                     'nodeSet': regexp_string[1:],  # 좌측 ',' 제거
                     'limit': limit,
                     'offset': offset,
                     'field': field_string}
    # for TEST
    current_app.logger.debug(rest_node.req)

    # Step 3-2. REST Call : get
    if not rest_node.get(url):
        return jsonify({'result': 'Fail'}), 500

    # Step 3-3. REST Call : Check Biz. Exception
    if rest_node.res['resultCode'] != OpmmResultCode.EM0000:
        rest_node.res['nodeList'] = []
        rest_node.res['totalCnt'] = 0
        return jsonify(rest_node.res)

    rest_res = rest_node.res
    total_cnt = rest_node.res['totalCnt']
    node_list += rest_node.res['nodeList']

    while total_cnt > offset:

        offset += limit
        rest_node.req['offset'] = offset

        # Step 3-2. REST Call : get
        if not rest_node.get(url):
            return jsonify({'result': 'Fail'}), 500

        # Step 3-3. REST Call : Check Biz. Exception
        # NOT_FOUND
        if rest_node.res['resultCode'] == OpmmResultCode.EM1002:
            break

        rest_res = rest_node.res
        node_list += rest_node.res['nodeList']

    rest_res['nodeList'] = node_list

    # Step 3-4. REST Call : Output Process

    # Step 99. Response Data Formatting
    return jsonify(rest_res)


@bp.route('/dtl', methods=['POST'])
@login_required
def request_node():
    """
    노드 정보를 가져온다.
    :return:
    """
    # Step 1. Request Data Parsing.
    node_session_id = request.form.get('node_session_id')

    if node_session_id is None or node_session_id == '':  # error
        error_map = {'resultCode': OpmmResultCode.EM0999,
                     'resultMsg': 'Do not found. nodeSessionId is Empty.'}
        return Render.render_template('node/node_dtl.html',
                                      request_params=request.form,
                                      response=error_map)

    # Step 2. Request Data Validation.
    #   - None.

    # Step 3. REST Call
    rest_node = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/nodes/' + node_session_id

    # Step 3-2. REST Call : get
    if not rest_node.get(url):
        error_map = {'resultCode': OpmmResultCode.EM0999,
                     'resultMsg': 'Internal Error'}
        return Render.render_template('node/node_lst.html',
                                      request_params=request.form,
                                      response=error_map)

    # Step 3-3. REST Call : Check Biz. Exception
    # NOT NORMAL
    if rest_node.res['resultCode'] != OpmmResultCode.EM0000:
        current_app.logger.error("[{0}] {1}".format(rest_node.res['resultCode'], rest_node.res['resultMsg']))
        return Render.render_template('node/node_lst.html',
                                      request_params=request.form,
                                      response=rest_node.res)

    # Step 3-4. REST Call : Output Process

    # Step 99. Response Data Formatting
    return Render.render_template('node/node_dtl.html',
                                  request_params=request.form,
                                  response=rest_node.res)


@bp.route('/p_node', methods=['GET'])
@login_required
def render_search_user_page():
    """
    노드 조회 Popup 화면
    :return:
    """
    return Render.render_template('node/popup/popup_node_lst.html')


@bp.route('/p_node_tag', methods=['GET'])
@login_required
def render_node_tagview_page():
    """
    모의실행 Popup 화면
    :return:
    """
    return Render.render_template('node/popup/popup_node_tagview.html')

@bp.route('/xl_export', methods=['POST'])
@login_required
def request_exec_xl_export():
    """
    실행결과 엑셀 내보내기
    :return:
    """
    # Step 1. Request Data Parsing.
    current_date = datetime.now().strftime("%Y%m%d")
    export_path = current_app.config['OPME_FILE_EXPORT_PATH'] + '/' + current_date

    # 실행 정보 조회 조건 받기
    base_condition = _parse_json_field(request.form.get('base_condition'))
    nodeset_svr_condition = _parse_json_field(request.form.get('nodeset_svr_condition'))
    nodeset_tag_condition = _parse_json_field(request.form.get('nodeset_tag_condition'))
    extra_fields = request.form.get('extra_fields') or ""
    save_filename = (request.form.get('uuid') or "").strip()

    # Create File Download Directory(from OPMM)
    if os.path.isdir(export_path) is False:
        os.mkdir(export_path)

    # 반복 조회 수행 시 offset, limit 필요.
    offset = 0
    limit = 100
    total_cnt = int(request.form.get('total_cnt'))

    sse_enable = current_app.config['SSE_ENABLE']

    # Step 2. Request Data Validation.
    #   - None.
    class ExecutionExcelExport(ExcelExport):

        def __init__(self):
            super().__init__()

            # Code dictionary
            self._column_definition = {
                # key: ColumnInfo(col_letter, title, align, decode_yn, code_dict)
                # Col 'A' : Empty
                'WP.customer_nm'  : self.ColumnInfo('B', '고객사', align='left'),
                'nodeSessionId': self.ColumnInfo('C', '노드세션 ID', align='left'),
                'hostname'     : self.ColumnInfo('D', 'Hostname'),
                'WP.sys_operator1_nm'     : self.ColumnInfo('E', '운영자'),
                'osType'       : self.ColumnInfo('F', 'OS 종류'),
                'osName'       : self.ColumnInfo('G', 'OS 이름'),
                'osVer'        : self.ColumnInfo('H', 'OS 버전'),
                'remoteAddr'   : self.ColumnInfo('I', 'IP 주소', align='left'),
                'heartbeat'    : self.ColumnInfo('J', 'Heart Beat'),
                'pastSession'  : self.ColumnInfo('K', 'Conflict'),
                'WP.service_nm'  : self.ColumnInfo('L', '서비스명'),
                'WP.env'   : self.ColumnInfo('M', '운영여부', align='left'),
                'cspTagName'   : self.ColumnInfo('N', 'Name', align='left'),
                'CT.cz-project'   : self.ColumnInfo('O', 'cz-project', align='left'),
                'CT.cz-owner'   : self.ColumnInfo('P', 'cz-owner', align='left'),
                'CT.cz-org'   : self.ColumnInfo('Q', 'cz-org', align='left'),
                'CT.cz-stage'   : self.ColumnInfo('R', 'cz-stage', align='left'),
                'CT.cz-appl'   : self.ColumnInfo('S', 'cz-appl', align='left'),
            }

    # OpenPyXL
    wb = Workbook()
    ws = wb.active
    ws.title = "OPMATE"
    ex = ExecutionExcelExport()

    # Set table header columns
    ex.set_header_column(ws, fixed_row=4)

    # Step 3. REST Call
    rest_node = RestClient()

    # Step 3-1. REST Call : Input Binding.
    url = '/nodes'
    field_string = "nodeSessionId,agentVer,osType,osName,osVer,cspResourceId,remoteAddr,hostname,pastSession,heartbeat,cspTagName,CT.Name,CT.cz-project,CT.cz-stage,CT.cz-org,CT.cz-owner,CT.cz-appl,WP.customer_nm,WP.sys_operator1_nm,WP.service_nm,WP.env"
    query_string = ''

    # 문자열 구성
    def build_query(d):
        # k=v 형태, 값이 비어있으면 제외
        parts = []
        for k, v in (d or {}).items():
            if v is None or str(v) == "":
                continue
            parts.append(f"{k}={v}")
        return ",".join(parts)
    def build_regexp(d):
        # k:"v" 형태
        parts = []
        for k, v in (d or {}).items():
            if v is None or str(v) == "":
                continue
            parts.append(f'{k}:"{v}"')
        return ",".join(parts)
    query_string  = build_query(base_condition)
    regexp_string = ",".join(filter(None, [
        build_regexp(nodeset_svr_condition),
        build_regexp(nodeset_tag_condition)
    ]))

    rest_node.req = {
        'query': query_string,
        'nodeSet': regexp_string,
        'field': field_string,
        'limit': limit,
        'offset': offset
    }

    # 루프 전
    current_cnt = 0
    while total_cnt > offset:

        rest_node.req['offset'] = offset

        if not rest_node.get(url):
            return jsonify({'result': 'Fail'}), 500

        if rest_node.res['resultCode'] != OpmmResultCode.EM0000:
            current_app.logger.error(
                "[{0}] {1}".format(rest_node.res['resultCode'], rest_node.res['resultMsg']))
            abort(500)

        for n in range(rest_node.res['currentCnt']):
            node = rest_node.res['nodeList'][n]
            # --- 추가: 응답값을 엑셀 헤더 키에 맞게 평탄화 ---
            # ExtTagList (origin=WP)
            node['WP.customer_nm'] = next((t.get('val', '') for t in node.get('extTagList', []) if
                                           t.get('origin') == 'WP' and t.get('key') == 'customer_nm'), '')
            node['WP.sys_operator1_nm'] = next((t.get('val', '') for t in node.get('extTagList', []) if
                                                t.get('origin') == 'WP' and t.get('key') == 'sys_operator1_nm'), '')
            node['WP.service_nm'] = next((t.get('val', '') for t in node.get('extTagList', []) if
                                          t.get('origin') == 'WP' and t.get('key') == 'service_nm'), '')
            node['WP.env'] = next((t.get('val', '') for t in node.get('extTagList', []) if
                                   t.get('origin') == 'WP' and t.get('key') == 'env'), '')

            # CSPTagList (CT)
            node['CT.cz-project'] = next(
                (t.get('val', '') for t in node.get('cspTagList', []) if t.get('key') == 'cz-project'), '')
            node['CT.cz-owner'] = next(
                (t.get('val', '') for t in node.get('cspTagList', []) if t.get('key') == 'cz-owner'), '')
            node['CT.cz-org'] = next((t.get('val', '') for t in node.get('cspTagList', []) if t.get('key') == 'cz-org'),
                                     '')
            node['CT.cz-stage'] = next(
                (t.get('val', '') for t in node.get('cspTagList', []) if t.get('key') == 'cz-stage'), '')
            node['CT.cz-appl'] = next(
                (t.get('val', '') for t in node.get('cspTagList', []) if t.get('key') == 'cz-appl'), '')

            ex.set_data_column(ws, node, n + offset + 4)

            for col_idx in range(1, ws.max_column + 1):
                cell_ref = ws.cell(row=n + offset + 4, column=col_idx)
                cell_value = cell_ref.value

                # if col_idx in [13, 14]:
                #     cell_ref.value = f"'{cell_value}"
                cell_ref.number_format = '@'

            current_cnt += 1

            if sse_enable == 'yes':
                if sse_dict.get(save_filename) is None:
                    current_app.logger.error("Aborted excel file export.")
                    delete_file = os.path.join(export_path, save_filename)
                    try:
                        if os.path.isfile(delete_file):
                            os.remove(delete_file)
                    except OSError:
                        abort(500)
                    abort(500)

                sse_dict[save_filename].put(u'%s' % int(current_cnt / total_cnt * 100))

        offset += limit

    # 열 너비 맞춤
    ex.autofit_cell_size(ws)

    # Last Summary add
    summary = "총 건수 : " + str(total_cnt)
    ex.set_additional_info_row(ws, summary)

    # 파일명 지정, 저장 처리
    file_name = export_path + "/" + save_filename
    wb.save(filename=file_name)

    '''
    # 처리 변경 (send_file 아닌 방법)
    def generate():
        try:
            with open(file_name, "rb") as f:
                yield from f
        except Exception as e:
            print("Exception : {0}".format(e))
        print('delete_file start')
        os.remove(file_name)
        print('delete_file end')
    r = current_app.response_class(generate(), mimetype='application/octet-stream')
    r.headers.set('Content-Disposition', 'attachment', filename=file_name)
    return r
    '''
    return send_file(os.path.join(export_path, file_name),
                     download_name=current_date + '/' + file_name,
                     as_attachment=True)

@bp.route('/del_tmp_file', methods=['POST'])
@login_required
def delete_tmp_file():
    """
    실행결과 엑셀 내보내기 완료 후, 호출하여 임시 파일 삭제.
    :return:
    """
    req = request.get_json()
    date = req['date']
    file_name = req['file_name']
    export_path = current_app.config['OPME_FILE_EXPORT_PATH'] + "/" + date
    delete_file = os.path.join(export_path, file_name)

    try:
        if os.path.isfile(delete_file):
            os.remove(delete_file)

    except OSError:
        return jsonify({'result': 'Fail'}), 500

    return jsonify({'result': 'Success'})


@bp.route('/sse/<uuid>', methods=['GET'])
@login_required
def export_sse(uuid):
    if not sse_dict.get(uuid) is None:
        return None
    return Response(stream_with_context(event_stream(uuid)),
                    mimetype="text/event-stream",
                    headers={"X-Accel-Buffering": "no"})


def event_stream(uuid):
    try:
        sse_dict[uuid] = Queue()
        progress_percent = 0

        while int(progress_percent) < 100:
            progress_percent = sse_dict[uuid].get()
            yield 'data: %s\n\n' % progress_percent
    finally:
        if not sse_dict.get(uuid) is None:
            del(sse_dict[uuid])